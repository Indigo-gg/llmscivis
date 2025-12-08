# 调用embedding_v3_1.py中的函数，进行检索，输入和输出与retiever.py一致

from typing import Dict, List, Optional
import sys
import os
import json
import pandas as pd
from RAG.embedding_v3_1 import index, load_faiss_index

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 导入embedding_v3_1.py中的函数
from RAG.embedding_v3_1 import search_code_optimized, search_code_optimized_with_stages, analyze_query as embedding_analyze_query, rerank_results
from RAG.embedding_v3_1 import embed_text, mongo_manager, index, embedding_dim
from RAG.embedding_v3_1 import K, Similarity_Threshold
from config.app_config import app_config
from config.ollama_config import ollama_config
from llm_agent.prompt_agent import analyze_query as prompt_analyze_query
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import gc
import atexit

# 注册清理函数，在程序退出时执行
def _cleanup_on_exit():
    """为openpyxl配置清理底每，防止atexit错误"""
    import warnings
    warnings.filterwarnings('ignore')
    gc.collect()
    
atexit.register(_cleanup_on_exit)


def check_mongodb_connection():
    """
    检查MongoDB连接状态
    
    Returns:
        bool: 连接成功返回True，失败返回False
    """
    try:
        import pymongo
        client = pymongo.MongoClient('localhost', 27017, serverSelectionTimeoutMS=5000)
        # 触发实际连接
        client.admin.command('ping')
        print("✓ MongoDB 连接成功")
        client.close()
        return True
    except Exception as e:
        print(f"✗ MongoDB 连接失败: {e}")
        print("\n请确保 MongoDB 服务已启动。启动方式:")
        print("  1. 打开新的终端窗口")
        print("  2. 运行命令: mongod")
        print("  3. 等待 MongoDB 服务启动完成后，再运行本脚本\n")
        return False


def save_retrieval_results_to_excel(output_file, benchmark_prompts, all_raw_results, all_reranked_results):
    """
    将初筛结果和重排序结果保存到Excel表格中
    
    Args:
        output_file: 输出Excel文件路径
        benchmark_prompts: 原始benchmark提示词列表
        all_raw_results: 所有初筛结果（第一步）
        all_reranked_results: 所有重排序结果（第二步）
    """
    print(f"[DEBUG] 开始保存Excel文件: {output_file}")
    print(f"[DEBUG] all_raw_results 结构: {type(all_raw_results)}, 长度: {len(all_raw_results) if all_raw_results else 0}")
    print(f"[DEBUG] all_reranked_results 结构: {type(all_reranked_results)}, 长度: {len(all_reranked_results) if all_reranked_results else 0}")
    
    # 检查是否有数据需要保存
    if not all_raw_results or len(all_raw_results) == 0:
        print(f"[WARNING] 没有初筛结果，跳过Excel保存")
        return
    
    try:
        wb = Workbook()
        
        # 创建第一个sheet：初筛结果
        ws_raw = wb.active
        ws_raw.title = "Stage 1 - Initial Retrieval"
        
        # 创建第二个sheet：重排序结果
        ws_reranked = wb.create_sheet("Stage 2 - Reranked Results")
        
        # ========== 第一步：初筛结果表 ==========
        raw_headers = [
            "Query Index",
            "Benchmark Prompt",
            "Query Description",
            "Result Index",
            "File Path",
            "FAISS ID",
            "FAISS Similarity",
            "Meta Title",
            "Meta Description",
            "VTK.js Modules",
            "Input File",
            "Output File"
        ]
        
        ws_raw.append(raw_headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws_raw[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 填充初筛结果数据
        row_idx = 2
        # all_raw_results是嵌套列表：第一层是不同的search()调用（不同的benchmark），第二层是查询，第三层是结果
        for search_call_idx, search_call_raw_results in enumerate(all_raw_results):
            # search_call_raw_results 是一个查询列表的结果集合
            for query_idx, results_list in enumerate(search_call_raw_results):
                # 如果有多个benchmark提示词对应多个search()调用，这里需要计算正确的benchmark_prompt索引
                benchmark_prompt_idx = search_call_idx * len(search_call_raw_results) + query_idx
                benchmark_prompt = benchmark_prompts[benchmark_prompt_idx] if benchmark_prompt_idx < len(benchmark_prompts) else "N/A"
                
                # 处理嵌套的结果列表
                if not results_list or (isinstance(results_list, list) and len(results_list) == 0):
                    ws_raw.append([
                        benchmark_prompt_idx + 1,
                        benchmark_prompt[:100] if isinstance(benchmark_prompt, str) else benchmark_prompt,
                        "No query",
                        "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"
                    ])
                else:
                    # results_list是单个查询的所有结果列表
                    if not isinstance(results_list, list):
                        # 处理异常情况：如果不是列表，跳过
                        continue
                        
                    for result_idx, result in enumerate(results_list):
                        if not isinstance(result, dict):
                            continue
                            
                        meta_info = result.get('meta_info', {})
                        query_desc = result.get('query_description', 'N/A')
                        
                        ws_raw.append([
                            benchmark_prompt_idx + 1,
                            benchmark_prompt[:100] if isinstance(benchmark_prompt, str) else benchmark_prompt,
                            query_desc if isinstance(query_desc, str) else query_desc,
                            result_idx + 1,
                            result.get('file_path', 'N/A'),
                            result.get('faiss_id', 'N/A'),
                            round(result.get('faiss_similarity', 0), 4),
                            meta_info.get('title', 'N/A'),
                            meta_info.get('description', 'N/A')[:100],
                            ', '.join(meta_info.get('vtkjs_modules', [])) if meta_info.get('vtkjs_modules') else 'N/A',
                            result.get('input_file', 'N/A'),
                            result.get('output_file', 'N/A')
                        ])
        
        # 调整列宽
        column_widths = [12, 30, 30, 12, 40, 12, 15, 20, 30, 30, 20, 20]
        for idx, width in enumerate(column_widths, 1):
            col_letter = chr(64 + idx)
            ws_raw.column_dimensions[col_letter].width = width
        
        # ========== 第二步：重排序结果表 ==========
        reranked_headers = [
            "Query Index",
            "Benchmark Prompt",
            "Query Description",
            "Result Index",
            "File Path",
            "FAISS ID",
            "FAISS Similarity",
            "Rerank Score",
            "Meta Title",
            "Meta Description",
            "VTK.js Modules",
            "Input File",
            "Output File"
        ]
        
        ws_reranked.append(reranked_headers)
        
        # 设置表头样式
        for cell in ws_reranked[1]:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 填充重排序结果数据
        row_idx = 2
        # all_reranked_results是嵌套列表：第一层是不同的search()调用（不同的benchmark），第二层是查询，第三层是结果
        for search_call_idx, search_call_reranked_results in enumerate(all_reranked_results):
            # search_call_reranked_results 是一个查询列表的结果集合
            for query_idx, results_list in enumerate(search_call_reranked_results):
                # 如果有多个benchmark提示词对应多个search()调用，这里需要计算正确的benchmark_prompt索引
                benchmark_prompt_idx = search_call_idx * len(search_call_reranked_results) + query_idx
                benchmark_prompt = benchmark_prompts[benchmark_prompt_idx] if benchmark_prompt_idx < len(benchmark_prompts) else "N/A"
                
                # 处理嵌套的结果列表
                if not results_list or (isinstance(results_list, list) and len(results_list) == 0):
                    ws_reranked.append([
                        benchmark_prompt_idx + 1,
                        benchmark_prompt[:100] if isinstance(benchmark_prompt, str) else benchmark_prompt,
                        "No query",
                        "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"
                    ])
                else:
                    # results_list是单个查询的所有结果列表
                    if not isinstance(results_list, list):
                        # 处理异常情况：如果不是列表，跳过
                        continue
                        
                    for result_idx, result in enumerate(results_list):
                        if not isinstance(result, dict):
                            continue
                            
                        meta_info = result.get('meta_info', {})
                        query_desc = result.get('query_description', 'N/A')
                        
                        ws_reranked.append([
                            benchmark_prompt_idx + 1,
                            benchmark_prompt[:100] if isinstance(benchmark_prompt, str) else benchmark_prompt,
                            query_desc  if isinstance(query_desc, str) else query_desc,
                            result_idx + 1,
                            result.get('file_path', 'N/A'),
                            result.get('faiss_id', 'N/A'),
                            round(result.get('faiss_similarity', 0), 4),
                            round(result.get('rerank_score', 0), 4),
                            meta_info.get('title', 'N/A'),
                            meta_info.get('description', 'N/A')[:100],
                            ', '.join(meta_info.get('vtkjs_modules', [])) if meta_info.get('vtkjs_modules') else 'N/A',
                            result.get('input_file', 'N/A'),
                            result.get('output_file', 'N/A')
                        ])
        
        # 调整列宽
        column_widths = [12, 30, 30, 12, 40, 12, 15, 15, 20, 30, 30, 20, 20]
        for idx, width in enumerate(column_widths, 1):
            col_letter = chr(64 + idx)
            ws_reranked.column_dimensions[col_letter].width = width
        
        # 保存Excel文件
        wb.save(output_file)
        wb.close()  # 显式关闭工作簿
        print(f"\n✓ 检索结果已保存到: {output_file}")
        print(f"  - Sheet 1: Stage 1 - Initial Retrieval (初筛结果)")
        print(f"  - Sheet 2: Stage 2 - Reranked Results (重排序结果)")
        print(f"[DEBUG] Excel文件保存成功")
        
    except Exception as e:
        print(f"[DEBUG] 异常信息: {type(e).__name__}: {e}")
        print(f"✗ 保存检索结果到Excel时出错: {e}")
        import traceback
        traceback.print_exc()
        # 尝试关闭工作簿
        try:
            wb.close()
        except:
            pass
    
    finally:
        # 确保总是进行资源清理
        try:
            import gc
            gc.collect()
        except:
            pass


def search_with_stages(query: str, k: int = 4, similarity_threshold: float = 0.1):
    """
    执行两阶段搜索：初筛和重排序，并返回两个阶段的结果
    这是对 embedding_v3_1.search_code_optimized_with_stages() 的包装函数。
    
    Args:
        query: 查询文本
        k: 最终返回的结果数
        similarity_threshold: 相似度阈值
        
    Returns:
        tuple: (raw_results, reranked_results)
            - raw_results: 初筛结果（第一阶段 - FAISS语义搜索）
            - reranked_results: 重排序结果（第二阶段 - 重排）
    """
    # 直接调用 embedding_v3_1 中的新函数
    return search_code_optimized_with_stages(query, k, similarity_threshold, recall_k_multiplier=2)

class VTKSearcherV2:
    def __init__(self):
        """
        初始化 VTKSearcherV2 类。
        与 VTKSearcher 类似，但使用 embedding_v3_1.py 中的检索逻辑。
        """
        # 这里不需要加载FAISS索引，因为embedding_v3_1.py已经处理了
        self.last_retrieval_metadata = []
        self.raw_results_history = []  # 存储所有初筛结果
        self.reranked_results_history = []  # 存储所有重排序结果

    def search(self, query:str, query_list: List) -> str:
        """
        使用两阶段搜索执行检索，优量INFORMATION
        呜时记录初筛和重排序的结果。
        """
        # 检查FAISS索引是否已加载
        if index.ntotal == 0:
            print("检测到FAISS索引未加载，正在尝试加载...")
            load_faiss_index("faiss_index.index")
            
        # 使用两阶段搜索
        # 注意：保存为嵌套列表结构，便于Excel导出
        # 每个查询对应一个结果列表
        query_raw_results_list = []  # 按查询分别保存初筛结果
        query_reranked_results_list = []  # 按查询分别保存重排序结果
        all_reranked_for_context = []  # 用于上下文构建的所有重排序结果
        
        for query_item in query_list:
            # 从每个查询项中提取描述文本
            query_text = query_item['description']
            # 调用两阶段搜索函数
            print(f"Processing query: {query_text}")
            raw_results, reranked_results = search_with_stages(query_text, k=4, similarity_threshold=0.1)
            
            # 记录查询描述到两个结果集合中
            for result in raw_results:
                result['query_description'] = query_text
            for result in reranked_results:
                result['query_description'] = query_text
            
            # 按查询分别保存结果（嵌套列表结构）
            query_raw_results_list.append(raw_results)
            query_reranked_results_list.append(reranked_results)
            
            # 用于上下文构建的所有结果
            all_reranked_for_context.extend(reranked_results)
        
        # 收集最终的重排序结果（用于上下文构建）
        unique_results = {}
        for result in all_reranked_for_context:
            faiss_id = result.get("faiss_id")
            rerank_score = result.get("rerank_score", 0)
            
            if faiss_id not in unique_results or unique_results[faiss_id].get("rerank_score", 0) < rerank_score:
                unique_results[faiss_id] = result
        
        # 使用去重后的结果
        search_results = list(unique_results.values())
        search_results.sort(key=lambda x: x.get("rerank_score", 0), reverse=True)
        
        # 记录两个阶段的结果（嵌套列表结构）
        self.raw_results_history.append(query_raw_results_list)
        self.reranked_results_history.append(query_reranked_results_list)
        print(f"[DEBUG] 已保存到history: raw_results_history长度={len(self.raw_results_history)}, reranked_results_history长度={len(self.reranked_results_history)}")
        
        # Store retrieval metadata for frontend display
        self.last_retrieval_metadata = []
        for idx, result in enumerate(search_results[:10]):  # Top 10 results
            meta = result.get("meta_info", {})
            self.last_retrieval_metadata.append({
                "id": result.get("faiss_id", idx),
                "title": meta.get("title", f"Example {idx+1}"),
                "description": meta.get("description", "N/A")[:200],
                "relevance": result.get("rerank_score", 0.0)
            })
        
        # 构建上下文信息
        context_parts = []
        
        # 从检索结果中提取相关信息并格式化
        if search_results:
            for j, result in enumerate(search_results):
                meta = result.get("meta_info", {})
                code = result.get("code", "N/A")
                description = meta.get("description", "N/A")
                vtkjs_modules = meta.get("vtkjs_modules", "N/A")
                # 构建单个结果的上下文信息
                result_context = f"示例 {j+1}:\n"
                result_context += f"描述: {description}\n"
                result_context += f"VTK.js 模块: {vtkjs_modules}\n"
                result_context += f"代码:\n{code}\n"
                context_parts.append(result_context)
        else:
            context_parts.append("未找到相关示例")
        
        # 组合上下文片段
        if not context_parts:
            context = "No relevant VTK examples found matching the criteria."
        else:
            context = "\n--------------------------------------------------------------------------------------------------------------\n".join(context_parts)
        
        # 构建最终的提示
        final_prompt = f"""
    Generate only the HTML code without any additional text or markdown formatting.
    User Requirements:
    {query}

    Relevant VTK.js Examples:
    {context}
            """
        
        return final_prompt


def read_benchmark_prompts(input_file="res2.xlsx"):
    """
    读取Excel文件中的Benchmark prompt列，返回所有prompt的列表
    
    Args:
        input_file: 输入Excel文件路径
        
    Returns:
        list: 包含所有Benchmark prompt的列表
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(input_file, sheet_name='第二期实验数据')
        
        # 检查是否存在Benchmark prompt列
        if 'Benchmark prompt' not in df.columns:
            raise ValueError("Excel文件中必须包含'Benchmark prompt'列")
        
        benchmark_prompts_list = []
        
        # 处理每个Benchmark prompt
        for index, row in df.iterrows():
            benchmark_prompt = row['Benchmark prompt']
            
            if pd.isna(benchmark_prompt) or benchmark_prompt == '':
                print(f"跳过第{index+1}行：空的Benchmark prompt")
                benchmark_prompts_list.append("")  # 添加空字符串
                continue
                
            benchmark_prompts_list.append(benchmark_prompt)
        
        return benchmark_prompts_list
        
    except Exception as e:
        print(f"读取Benchmark prompt时出错: {e}")
        return []
def read_splited_prompts(input_file="res2.xlsx"):
    """
    读取Excel文件中的splited_prompt列，解析其中的JSON列表数据
    
    Args:
        input_file: 输入Excel文件路径
        
    Returns:
        list: 包含所有解析后的splited_prompt列表的列表
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(input_file, sheet_name='第二期实验数据')
        
        # 检查是否存在splited_prompt列
        if 'splited_prompt' not in df.columns:
            raise ValueError("Excel文件中必须包含'splited_prompt'列")
        
        splited_prompts_list = []
        
        # 处理每个splited_prompt
        for index, row in df.iterrows():
            splited_prompt_str = row['splited_prompt']
            
            if pd.isna(splited_prompt_str) or splited_prompt_str == '' or splited_prompt_str == '分析失败':
                print(f"跳过第{index+1}行：空的或无效的splited_prompt")
                splited_prompts_list.append([])  # 添加空列表
                continue
                
            try:
                # 解析JSON字符串
                splited_prompt = json.loads(splited_prompt_str)
                if isinstance(splited_prompt, list):
                    splited_prompts_list.append(splited_prompt)
                    print(f"第{index+1}行解析成功，包含{len(splited_prompt)}个元素")
                else:
                    print(f"第{index+1}行解析结果不是列表格式")
                    splited_prompts_list.append([])
            except json.JSONDecodeError as e:
                print(f"第{index+1}行JSON解析失败: {e}")
                splited_prompts_list.append([])
            except Exception as e:
                print(f"第{index+1}行处理出错: {e}")
                splited_prompts_list.append([])
        
        return splited_prompts_list
        
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return []


def process_splited_prompts_for_excel(input_file="res2.xlsx"):
    """
    读取splited_prompt数据并为RAG检索做准备
    """
    try:
        # 读取分割后的提示词
        splited_prompts = read_splited_prompts(input_file)
        
        if not splited_prompts:
            print("未读取到有效的splited_prompt数据")
            return []
        
        print(f"成功读取到{len(splited_prompts)}组分割后的提示词")
        
        # 处理数据以适应RAG的输入格式
        rag_input_data = []
        for i, prompt_list in enumerate(splited_prompts):
            if prompt_list and isinstance(prompt_list, list):
                # 确保每个元素都是字典且包含description字段
                valid_prompts = []
                for item in prompt_list:
                    if isinstance(item, dict) and 'description' in item:
                        valid_prompts.append(item)
                    elif isinstance(item, str):
                        # 如果是字符串，转换为标准格式
                        valid_prompts.append({'description': item})
                
                if valid_prompts:
                    rag_input_data.append(valid_prompts)
                    print(f"第{i+1}组提示词包含{len(valid_prompts)}个有效项")
                else:
                    rag_input_data.append([])
                    print(f"第{i+1}组没有有效提示词")
            else:
                rag_input_data.append([])
                print(f"第{i+1}组不是有效的提示词列表")
        
        return rag_input_data
        
    except Exception as e:
        print(f"处理splited_prompt数据时出错: {e}")
        return []


def get_data_from_excel(input_file="res2.xlsx"):
    """
    读取Excel文件中的Benchmark prompt和splited_prompt列，返回两个列表
    """
    benchmark_prompts = read_benchmark_prompts(input_file)
    splited_prompts = process_splited_prompts_for_excel(input_file)
    
    return benchmark_prompts, splited_prompts


def generate_splited_prompts_from_benchmarks(input_file="res2.xlsx", model_name=None, output_file=None):
    """
    读取Excel文件中的Benchmark prompt列，调用prompt_agent生成分割后的提示词
    并保存结果到指定文件
    
    Args:
        input_file: 输入Excel文件路径
        model_name: 使用的LLM模型名称，若为None则使用默认配置
        output_file: 输出文件路径，若为None则覆盖输入文件
        
    Returns:
        list: 包含所有生成的分割提示词的列表
    """
    if model_name is None:
        model_name = ollama_config.models_qwen["qwen3-plus"]
    
    if output_file is None:
        output_file = input_file
    
    try:
        # 读取Excel文件
        df = pd.read_excel(input_file, sheet_name='第二期实验数据')
        
        # 检查是否存在Benchmark prompt列
        if 'Benchmark prompt' not in df.columns:
            raise ValueError("Excel文件中必须包含'Benchmark prompt'列")
        
        # 确保splited_prompt列存在
        if 'splited_prompt' not in df.columns:
            df['splited_prompt'] = ''
        if 'time_spend_prompt' not in df.columns:
            df['time_spend_prompt'] = ''
        
        # 存储所有生成的分割提示词
        all_splited_prompts = []
        
        # 处理每个Benchmark prompt
        for index, row in df.iterrows():
            benchmark_prompt = row['Benchmark prompt']
            
            if pd.isna(benchmark_prompt) or benchmark_prompt == '':
                print(f"跳过第{index+1}行：空的Benchmark prompt")
                df.at[index, 'splited_prompt'] = ''
                df.at[index, 'time_spend_prompt'] = ''
                all_splited_prompts.append([])
                continue
            
            print(f"正在处理第{index+1}行的Benchmark prompt...")
            print(f"原始提示词: {benchmark_prompt}")
            
            # 记录开始时间
            start_time = time.time()
            
            try:
                # 调用prompt_agent中的analyze_query函数（使用导入别名以避免冲突）
                result = prompt_analyze_query(benchmark_prompt, model_name)
                
                # 记录结束时间
                end_time = time.time()
                time_spent = end_time - start_time
                
                if result is not None:
                    # 将结果转换为JSON字符串保存
                    df.at[index, 'splited_prompt'] = json.dumps(result, ensure_ascii=False, indent=2)
                    df.at[index, 'time_spend_prompt'] = f"{time_spent:.2f}"
                    all_splited_prompts.append(result)
                    print(f"第{index+1}行处理完成，耗时: {time_spent:.2f}秒")
                    print(f"生成{len(result)}个分割提示词\n")
                else:
                    df.at[index, 'splited_prompt'] = '分析失败'
                    df.at[index, 'time_spend_prompt'] = f"{time_spent:.2f}"
                    all_splited_prompts.append([])
                    print(f"第{index+1}行分析失败，耗时: {time_spent:.2f}秒\n")
                    
            except Exception as e:
                end_time = time.time()
                time_spent = end_time - start_time
                
                df.at[index, 'splited_prompt'] = f"处理出错: {str(e)}"
                df.at[index, 'time_spend_prompt'] = f"{time_spent:.2f}"
                all_splited_prompts.append([])
                print(f"第{index+1}行处理出错: {e}，耗时: {time_spent:.2f}秒\n")
        
        # 保存结果到Excel文件
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='第二期实验数据', index=False)
            print(f"所有结果已保存到 {output_file}")
        except Exception as e:
            print(f"保存文件时出错: {e}")
            # 尝试另存为新文件
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            df.to_excel(backup_file, sheet_name='第二期实验数据', index=False)
            print(f"已保存到备份文件: {backup_file}")
        
        return all_splited_prompts
        
    except Exception as e:
        print(f"处理过程中出错: {e}")
        # 如果Excel保存失败，尝试保存为JSON作为备选
        try:
            json_backup = output_file.replace('.xlsx', '_backup.json')
            with open(json_backup, 'w', encoding='utf-8') as f:
                json.dump(all_splited_prompts, f, ensure_ascii=False, indent=2)
            print(f"已保存到JSON备份文件: {json_backup}")
        except Exception as backup_e:
            print(f"备份保存也失败了: {backup_e}")
        return []

if __name__ == "__main__":
    # 检查MongoDB连接
    print("="*60)
    print("[初始化检查]")
    print("="*60)
    if not check_mongodb_connection():
        print("\n程序终止，请先启动MongoDB服务")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("[开始处理]")
    print("="*60 + "\n")
        
    # 初始化搜索器
    searcher = VTKSearcherV2()
    
    # 文件路径
    excel_path = "D://Pcode//LLM4VIS//llmscivis//data//recoreds//res2.xlsx"
    benchmark_prompts,splited_queries=get_data_from_excel(excel_path)

    # 创建结果文件
    output_file = "retrieval_results_12_2.json"
    
    # 准备存储所有结果的列表
    all_results = []
    splited_queries=generate_splited_prompts_from_benchmarks(excel_path,output_file="splited_queries_12_2.xlsx")
    print(f"[DEBUG] 生成了 {len(splited_queries)} 组查询")
    print(f"[DEBUG] searcher.raw_results_history 初始长度: {len(searcher.raw_results_history)}")
    
    for i, query_list in enumerate(splited_queries):
        if not query_list:
            print(f"跳过第{i+1}行：空的或无效的splited_prompt")
            all_results.append({
                "index": i+1,
                "status": "skipped",
                "reason": "空的或无效的splited_prompt"
            })
            continue
        
        # 获取对应的原始查询
        original_query = benchmark_prompts[i] if i < len(benchmark_prompts) else "N/A"
        print(f"\n处理第{i+1}行，原始查询: {original_query}")
        
        # 进行搜索
        final_prompt = searcher.search(original_query, query_list)
        
        # 输出结果到控制台
        print(f"最终提示:\n{final_prompt}\n")
        
        # 保存结果到列表
        result_entry = {
            "index": i+1,
            "original_query": original_query,
            "splited_queries": query_list,
            "final_prompt": final_prompt,
            "status": "success"
        }
        all_results.append(result_entry)
    
    # 将最终的RAG检索结果写入JSON文件
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_results, f, ensure_ascii=False, indent=2)
        print(f"\n所有结果已保存到 {output_file} 文件中")
    except Exception as e:
        print(f"保存结果到JSON文件时出错: {e}")
    
    # 保存检索结果（初筛和重排序）到Excel
    retrieval_excel_file = "retrieval_results_detailed_12_2_15.xlsx"
    print(f"[DEBUG] 保存前 searcher.raw_results_history 长度: {len(searcher.raw_results_history)}")
    print(f"[DEBUG] 保存前 searcher.reranked_results_history 长度: {len(searcher.reranked_results_history)}")
    if searcher.raw_results_history:
        print(f"[DEBUG] raw_results_history[0] 长度: {len(searcher.raw_results_history[0])}")
    if searcher.reranked_results_history:
        print(f"[DEBUG] reranked_results_history[0] 长度: {len(searcher.reranked_results_history[0])}")
    
    save_retrieval_results_to_excel(
        retrieval_excel_file,
        benchmark_prompts,
        searcher.raw_results_history,  # 初筛结果
        searcher.reranked_results_history  # 重排序结果
    )
    
    print("\n流程完成！")
    print(f"- 分割后的提示词已保存到 Excel 文件")
    print(f"- RAG检索结果已保存到 {output_file} 文件中")
    print(f"- 详细的检索结果已保存到 {retrieval_excel_file} 文件中")
   